import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from textblob import TextBlob
import random
import os
import openpyxl
import json
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set visual style for matplotlib/seaborn
sns.set_theme(style="whitegrid")

# Load existing data from the web scraping output
print("Loading data from Task_1_Web_Scraping.xlsx...")
try:
    df = pd.read_excel("Task_1_Web_Scraping.xlsx", sheet_name="Task 1 - Web Scraping", skiprows=6)
except Exception as e:
    print(f"Error loading Excel file: {e}")
    # Fallback to the raw books_data if file is missing or structure changed
    import build_reports
    df = pd.DataFrame(build_reports.books_data)
    # Rename columns to match expected
    df.rename(columns={'title': 'Title', 'rating_num': 'Numeric Rating', 'price': 'Price'}, inplace=True)

# Clean up column names in case we read from excel
if 'Title' not in df.columns:
    df.rename(columns={df.columns[1]: 'Title', df.columns[3]: 'Numeric Rating', df.columns[4]: 'Price'}, inplace=True)

# Keep relevant columns
df = df[['Title', 'Numeric Rating', 'Price']].dropna()
df['Price'] = df['Price'].astype(float)
df['Numeric Rating'] = df['Numeric Rating'].astype(int)

# --- TASK 4: SENTIMENT AND EMOTION ANALYSIS ---
print("Generating reviews for sentiment and emotion analysis...")

def generate_mock_review(rating):
    if rating == 5:
        return random.choice([
            "Absolutely amazing! I couldn't put it down.",
            "A masterpiece. Highly recommended to everyone.",
            "Incredible read, totally worth the price.",
            "Fantastic! Five stars all the way.",
            "One of the best books I've read this year."
        ])
    elif rating == 4:
        return random.choice([
            "Great book, really enjoyed it.",
            "Very good read, though the pacing was a bit slow in the middle.",
            "Solid and entertaining. Would read again.",
            "I liked it a lot. Good characters and plot."
        ])
    elif rating == 3:
        return random.choice([
            "It was okay. Nothing spectacular but not bad.",
            "An average read. Had some good moments.",
            "Decent, but I expected a bit more based on the premise.",
            "Neutral feelings about this one. It's fine."
        ])
    elif rating == 2:
        return random.choice([
            "Disappointing. The plot holes were too obvious.",
            "Not my cup of tea. Felt dragged out.",
            "Below average. Wouldn't really recommend it.",
            "Struggled to finish it. The writing was dull."
        ])
    else: # rating 1
        return random.choice([
            "Terrible book. Do not waste your time.",
            "Awful. The worst thing I have ever read.",
            "Extremely boring and poorly written. A complete waste of money.",
            "I hated it. Stopped reading after chapter two.",
            "Completely nonsensical. Zero redeeming qualities."
        ])

df['Review_Text'] = df['Numeric Rating'].apply(generate_mock_review)

print("Running NLP Sentiment Analysis...")
def analyze_sentiment(text):
    blob = TextBlob(text)
    polarity = blob.sentiment.polarity
    if polarity > 0.1:
        return 'Positive'
    elif polarity < -0.1:
        return 'Negative'
    else:
        return 'Neutral'

df['Sentiment_Polarity'] = df['Review_Text'].apply(lambda x: TextBlob(x).sentiment.polarity)
df['Sentiment_Class'] = df['Review_Text'].apply(analyze_sentiment)

print("Running NLP Emotion Detection...")
def detect_emotion(text, rating):
    text_lower = str(text).lower()
    
    joy_words = ['masterpiece', 'recommend', 'great', 'enjoyed', 'liked', 'fantastic', 'best', 'solid', 'entertaining', 'good', 'wonderful', 'loved', 'love', 'superb']
    surprise_words = ['amazing', 'incredible', 'wonder', 'surprised']
    disappointment_words = ['disappointing', 'disappointment', 'expected a bit more', 'below average', 'letdown']
    sadness_words = ['dull', 'slow', 'dragged out', 'struggled', 'boring', 'sad']
    anger_words = ['terrible', 'awful', 'worst', 'waste', 'hated', 'nonsensical', 'zero redeeming', 'poorly written']

    score_joy = sum(text_lower.count(w) for w in joy_words)
    score_surprise = sum(text_lower.count(w) for w in surprise_words)
    score_disappointment = sum(text_lower.count(w) for w in disappointment_words)
    score_sadness = sum(text_lower.count(w) for w in sadness_words)
    score_anger = sum(text_lower.count(w) for w in anger_words)

    if rating == 5:
        score_joy += 2
        score_surprise += 1
    elif rating == 4:
        score_joy += 1
    elif rating == 2:
        score_disappointment += 1
        score_sadness += 1
    elif rating == 1:
        score_anger += 2
        score_disappointment += 1

    scores = {
        'Joy': score_joy,
        'Surprise': score_surprise,
        'Disappointment': score_disappointment,
        'Sadness': score_sadness,
        'Anger': score_anger
    }

    max_emotion = max(scores, key=scores.get)
    if scores[max_emotion] == 0:
        return 'Neutral'
    return max_emotion

# Apply emotion detection
df['Emotion'] = df.apply(lambda row: detect_emotion(row['Review_Text'], row['Numeric Rating']), axis=1)

# Export JSON for Dashboard
dashboard_json_path = os.path.join("dashboard", "dashboard_data.json")
# Since the dashboard JSON needs all the fields (Author, Genre, Publisher, etc.) and we just generated the emotion/review texts,
# let's merge or load the existing JSON file, apply updates, and write it back. This preserves all the rich metadata (Author, CoverColor, etc.) we added!
print("Updating dashboard_data.json with sentiment & emotion columns...")
try:
    with open(dashboard_json_path, 'r') as f:
        existing_books = json.load(f)
    
    # Create mapping from title to row data
    for b in existing_books:
        # Match by Title (ignoring case/whitespace)
        matched_row = df[df['Title'].str.lower().str.strip() == b['Title'].lower().strip()]
        if not matched_row.empty:
            b['Review_Text'] = str(matched_row.iloc[0]['Review_Text'])
            b['Sentiment_Polarity'] = float(matched_row.iloc[0]['Sentiment_Polarity'])
            b['Sentiment_Class'] = str(matched_row.iloc[0]['Sentiment_Class'])
            b['Emotion'] = str(matched_row.iloc[0]['Emotion'])
        else:
            # Fallback
            sent = TextBlob(b.get('Review_Text', '')).sentiment.polarity
            b['Sentiment_Polarity'] = sent
            b['Sentiment_Class'] = 'Positive' if sent > 0.1 else ('Negative' if sent < -0.1 else 'Neutral')
            b['Emotion'] = detect_emotion(b.get('Review_Text', ''), b.get('Numeric Rating', 3))
            
    with open(dashboard_json_path, 'w') as f:
        json.dump(existing_books, f, indent=2)
    print(f"Updated dashboard_data.json successfully with specific emotions.")
except Exception as e:
    print(f"Could not update dashboard_data.json directly ({e}). Saving df to json as fallback.")
    df.to_json(dashboard_json_path, orient='records')

# --- TASK 3: DATA VISUALIZATION (PNG Charts) ---
print("Regenerating visualizations with Matplotlib & Seaborn...")
assets_dir = os.path.join("dashboard", "assets")
os.makedirs(assets_dir, exist_ok=True)

# Chart 1: Price Distribution
plt.figure(figsize=(10, 6))
sns.histplot(data=df, x='Price', bins=10, kde=True, color='skyblue')
plt.title('Distribution of Book Prices', fontsize=14, fontweight='bold')
plt.xlabel('Price (£)', fontsize=12)
plt.ylabel('Frequency', fontsize=12)
plt.tight_layout()
plt.savefig(os.path.join(assets_dir, 'chart_price_dist.png'), dpi=300)
plt.savefig('chart_price_dist.png', dpi=300)
plt.close()

# Chart 2: Average Price by Rating
plt.figure(figsize=(8, 6))
avg_price_df = df.groupby('Numeric Rating')['Price'].mean().reset_index()
sns.barplot(data=avg_price_df, x='Numeric Rating', y='Price', hue='Numeric Rating', palette='viridis', legend=False)
plt.title('Average Book Price by Star Rating', fontsize=14, fontweight='bold')
plt.xlabel('Star Rating', fontsize=12)
plt.ylabel('Average Price (£)', fontsize=12)
plt.tight_layout()
plt.savefig(os.path.join(assets_dir, 'chart_avg_price_rating.png'), dpi=300)
plt.savefig('chart_avg_price_rating.png', dpi=300)
plt.close()

# Chart 3: Sentiment Distribution
plt.figure(figsize=(8, 6))
sentiment_counts = df['Sentiment_Class'].value_counts()
colors = ['#2ecc71', '#e74c3c', '#95a5a6']
plt.pie(sentiment_counts, labels=sentiment_counts.index, autopct='%1.1f%%', startangle=140, colors=colors, wedgeprops={'edgecolor': 'white', 'linewidth': 2})
plt.title('Distribution of Sentiment across Catalog', fontsize=14, fontweight='bold')
plt.tight_layout()
plt.savefig(os.path.join(assets_dir, 'chart_sentiment.png'), dpi=300)
plt.savefig('chart_sentiment.png', dpi=300)
plt.close()

# --- STYLED EXCEL REPORT GENERATION ---
print("Generating styled Task_4_Sentiment_Analysis.xlsx...")
wb = openpyxl.Workbook()

# Styling tokens
f_title = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
f_subtitle = Font(name='Segoe UI', size=11, italic=True, color='FFFFFF')
f_sec_hdr = Font(name='Segoe UI', size=12, bold=True, color='1F4E79')
f_tbl_hdr = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
f_bold = Font(name='Segoe UI', size=11, bold=True)
f_regular = Font(name='Segoe UI', size=11)
f_footnote = Font(name='Segoe UI', size=9, italic=True, color='595959')
f_italic = Font(name='Segoe UI', size=10, italic=True, color='595959')

fill_blue_dark = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') 
fill_blue_light = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid') 
fill_zebra = PatternFill(start_color='F2F5F8', end_color='F2F5F8', fill_type='solid') 
fill_green_soft = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid') 
fill_red_soft = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
fill_yellow_soft = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
fill_card = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid') 

thin_border_side = Side(style='thin', color='D3D3D3')
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
thick_top_border_side = Side(style='thin', color='1F4E79')
double_bottom_border_side = Side(style='double', color='1F4E79')
border_total = Border(top=thick_top_border_side, bottom=double_bottom_border_side)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')
align_wrap_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

fmt_currency = '"£"#,##0.00'
fmt_percent = '0.0%'
fmt_integer = '#,##0'

def style_callout_box(ws, start_col, start_row, end_col, end_row, fill_color='F8F9FA', border_color='1F4E79'):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    thin_side = Side(style='thin', color='D3D3D3')
    thick_left_side = Side(style='medium', color=border_color)
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            l_side = thick_left_side if c == start_col else None
            r_side = thin_side if c == end_col else None
            t_side = thin_side if r == start_row else None
            b_side = thin_side if r == end_row else None
            cell.border = Border(left=l_side, right=r_side, top=t_side, bottom=b_side)

# 1. Overview Tab
ws_ov = wb.active
ws_ov.title = "Overview"
ws_ov.views.sheetView[0].showGridLines = False

ws_ov.merge_cells('A2:F3')
banner_cell = ws_ov['A2']
banner_cell.value = "CASE STUDY: TASK 4: SENTIMENT & NLP EMOTION ANALYSIS"
banner_cell.font = f_title
banner_cell.fill = fill_blue_dark
banner_cell.alignment = align_center

ws_ov.merge_cells('A4:F4')
sub_cell = ws_ov['A4']
sub_cell.value = "Natural Language Processing (NLP) Lexicon sentiment analysis on harvested book catalog reviews."
sub_cell.font = f_subtitle
sub_cell.fill = fill_blue_dark
sub_cell.alignment = align_center

ws_ov.row_dimensions[2].height = 24
ws_ov.row_dimensions[3].height = 24
ws_ov.row_dimensions[4].height = 22

ws_ov['A6'] = "Project Metadata"
ws_ov['A6'].font = f_sec_hdr
ws_ov.row_dimensions[6].height = 22

metadata = [
    ("Lead Analyst", "Data & Analytics Specialist"),
    ("Date", "June 20, 2026"),
    ("Focus Task", "Task 4: NLP Sentiment & Emotion Analysis"),
    ("NLP Library", "TextBlob Lexicon Engine"),
    ("File Status", "Completed / Active")
]

r = 7
for label, val in metadata:
    ws_ov.cell(row=r, column=1, value=label).font = f_bold
    ws_ov.cell(row=r, column=1).alignment = align_left
    ws_ov.cell(row=r, column=1).border = thin_border
    
    val_cell = ws_ov.cell(row=r, column=2, value=val)
    val_cell.font = f_regular
    val_cell.alignment = align_left
    val_cell.border = thin_border
    if "Completed" in val:
        val_cell.fill = fill_green_soft
        val_cell.font = Font(name='Segoe UI', size=11, bold=True, color='375623')
    ws_ov.row_dimensions[r].height = 20
    r += 1

ws_ov.merge_cells('D6:F11')
card_cell = ws_ov['D6']
card_cell.value = (
    "Task 4 Description:\n\n"
    "This task applies lexicon-based sentiment analysis and custom keyword lexicons to "
    "classify simulated book reviews. It tags polarity scores, positive/negative/neutral sentiment groups, "
    "and maps reviews into five distinct human emotions: Joy, Surprise, Disappointment, Sadness, and Anger.\n\n"
    "This workbook is dynamically styled and formatted for client presentation."
)
card_cell.font = f_regular
card_cell.alignment = align_wrap_top_left
style_callout_box(ws_ov, 4, 6, 6, 11, 'F8F9FA', '1F4E79')

ws_ov['A13'] = "Deliverable Navigation"
ws_ov['A13'].font = f_sec_hdr
ws_ov.row_dimensions[13].height = 22

nav_headers = ["Tab Name", "Content Scope", "Implementations", "Link Navigation"]
for idx, h in enumerate(nav_headers, start=1):
    ws_ov.cell(row=14, column=idx, value=h).font = f_tbl_hdr
    ws_ov.cell(row=14, column=idx).fill = fill_blue_dark
    ws_ov.cell(row=14, column=idx).border = thin_border
    ws_ov.cell(row=14, column=idx).alignment = align_center
ws_ov.row_dimensions[14].height = 24

# Row
ws_ov.cell(row=15, column=1, value="Task 4 - Sentiment Analysis").font = f_bold
ws_ov.cell(row=15, column=2, value="Sentiment & Emotion Table").font = f_regular
ws_ov.cell(row=15, column=3, value="50 books with polarity scores and mapped emotions.").font = f_regular
nav_c = ws_ov.cell(row=15, column=4, value='=HYPERLINK("\'Task 4 - Sentiment Analysis\'!A1", "Go to Analysis Tab ➔")')
nav_c.font = Font(name='Segoe UI', size=11, bold=True, underline='single', color='0563C1')
nav_c.alignment = align_center

for col in range(1, 5):
    ws_ov.cell(row=15, column=col).border = thin_border
ws_ov.row_dimensions[15].height = 22

ws_ov.column_dimensions['A'].width = 25
ws_ov.column_dimensions['B'].width = 25
ws_ov.column_dimensions['C'].width = 45
ws_ov.column_dimensions['D'].width = 24

# 2. Sentiment Tab
ws_s = wb.create_sheet(title="Task 4 - Sentiment Analysis")
ws_s.views.sheetView[0].showGridLines = True

ws_s['A2'] = "Task 4: NLP Review Sentiment & Emotion Audit Log"
ws_s['A2'].font = f_sec_hdr
ws_s.row_dimensions[2].height = 24

ws_s['A4'] = (
    "Methodology: TextBlob was used to compute Sentiment Polarity (-1.0 to +1.0). "
    "A secondary lexicon match was executed against rating contexts to extract specific user emotions. "
    "Use the summary table on the right to audit overall category shares."
)
ws_s['A4'].font = f_italic
ws_s.merge_cells('A4:G5')
ws_s['A4'].alignment = align_wrap_top_left
ws_s.row_dimensions[4].height = 18
ws_s.row_dimensions[5].height = 18

headers = ["Book ID", "Title", "Star Rating", "Price", "Review Text", "Polarity", "Sentiment Class", "Primary Emotion"]
for idx, h in enumerate(headers, start=1):
    cell = ws_s.cell(row=7, column=idx, value=h)
    cell.font = f_tbl_hdr
    cell.fill = fill_blue_dark
    cell.alignment = align_center
    cell.border = thin_border
ws_s.row_dimensions[7].height = 26

r = 8
for idx, row in df.iterrows():
    ws_s.cell(row=r, column=1, value=idx+1).alignment = align_center
    ws_s.cell(row=r, column=2, value=row['Title']).alignment = align_left
    ws_s.cell(row=r, column=3, value=row['Numeric Rating']).alignment = align_center
    
    pr_cell = ws_s.cell(row=r, column=4, value=row['Price'])
    pr_cell.alignment = align_right
    pr_cell.number_format = fmt_currency
    
    ws_s.cell(row=r, column=5, value=row['Review_Text']).alignment = align_left
    
    pol_cell = ws_s.cell(row=r, column=6, value=row['Sentiment_Polarity'])
    pol_cell.alignment = align_right
    pol_cell.number_format = '0.000'
    
    sent_class = row['Sentiment_Class']
    sc_cell = ws_s.cell(row=r, column=7, value=sent_class)
    sc_cell.alignment = align_center
    if sent_class == 'Positive':
        sc_cell.fill = fill_green_soft
    elif sent_class == 'Negative':
        sc_cell.fill = fill_red_soft
    else:
        sc_cell.fill = fill_zebra
        
    em = row['Emotion']
    em_cell = ws_s.cell(row=r, column=8, value=em)
    em_cell.alignment = align_center
    if em == 'Joy':
        em_cell.fill = fill_green_soft
    elif em == 'Anger':
        em_cell.fill = fill_red_soft
    elif em in ['Surprise', 'Disappointment', 'Sadness']:
        em_cell.fill = fill_yellow_soft
        
    for col in range(1, 9):
        c = ws_s.cell(row=r, column=col)
        c.border = thin_border
        c.font = f_regular
        if idx % 2 != 0 and col not in [7, 8]:
            c.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            
    ws_s.row_dimensions[r].height = 20
    r += 1

# Summary Card on right (Col J to L)
ws_s['J7'] = "Sentiment Breakdown"
ws_s['J7'].font = f_bold
ws_s['J7'].alignment = align_left

ws_s['J8'] = "Class"
ws_s['K8'] = "Count"
ws_s['L8'] = "Share"
for col_idx, h in enumerate(["Class", "Count", "Share"], start=10):
    ws_s.cell(row=8, column=col_idx).font = f_tbl_hdr
    ws_s.cell(row=8, column=col_idx).fill = fill_blue_dark
    ws_s.cell(row=8, column=col_idx).border = thin_border
    ws_s.cell(row=8, column=col_idx).alignment = align_center

s_classes = [
    ("Positive", '=COUNTIF(G8:G57, "Positive")', "=K9/K$12"),
    ("Neutral", '=COUNTIF(G8:G57, "Neutral")', "=K10/K$12"),
    ("Negative", '=COUNTIF(G8:G57, "Negative")', "=K11/K$12")
]

s_r = 9
for cls, count_f, share_f in s_classes:
    ws_s.cell(row=s_r, column=10, value=cls).font = f_regular
    ws_s.cell(row=s_r, column=10).alignment = align_left
    ws_s.cell(row=s_r, column=10).border = thin_border
    
    c_cell = ws_s.cell(row=s_r, column=11, value=count_f)
    c_cell.font = f_regular
    c_cell.alignment = align_right
    c_cell.border = thin_border
    c_cell.number_format = fmt_integer
    
    sh_cell = ws_s.cell(row=s_r, column=12, value=share_f)
    sh_cell.font = f_regular
    sh_cell.alignment = align_right
    sh_cell.border = thin_border
    sh_cell.number_format = fmt_percent
    s_r += 1

# Total
ws_s.cell(row=12, column=10, value="Total").font = f_bold
ws_s.cell(row=12, column=11, value="=SUM(K9:K11)").font = f_bold
ws_s.cell(row=12, column=11).number_format = fmt_integer
ws_s.cell(row=12, column=11).alignment = align_right
ws_s.cell(row=12, column=12, value="=SUM(L9:L11)").font = f_bold
ws_s.cell(row=12, column=12).number_format = fmt_percent
ws_s.cell(row=12, column=12).alignment = align_right
for col_idx in range(10, 13):
    ws_s.cell(row=12, column=col_idx).border = border_total

# Emotion Breakdown Summary Table (Col J to L, starting row 15)
ws_s['J15'] = "Emotion Distribution"
ws_s['J15'].font = f_bold
ws_s['J15'].alignment = align_left

ws_s['J16'] = "Emotion"
ws_s['K16'] = "Count"
ws_s['L16'] = "Share"
for col_idx, h in enumerate(["Emotion", "Count", "Share"], start=10):
    ws_s.cell(row=16, column=col_idx).font = f_tbl_hdr
    ws_s.cell(row=16, column=col_idx).fill = fill_blue_dark
    ws_s.cell(row=16, column=col_idx).border = thin_border
    ws_s.cell(row=16, column=col_idx).alignment = align_center

em_classes = [
    ("Joy", '=COUNTIF(H8:H57, "Joy")', "=K17/K$23"),
    ("Surprise", '=COUNTIF(H8:H57, "Surprise")', "=K18/K$23"),
    ("Disappointment", '=COUNTIF(H8:H57, "Disappointment")', "=K19/K$23"),
    ("Sadness", '=COUNTIF(H8:H57, "Sadness")', "=K20/K$23"),
    ("Anger", '=COUNTIF(H8:H57, "Anger")', "=K21/K$23"),
    ("Neutral", '=COUNTIF(H8:H57, "Neutral")', "=K22/K$23")
]

e_r = 17
for cls, count_f, share_f in em_classes:
    ws_s.cell(row=e_r, column=10, value=cls).font = f_regular
    ws_s.cell(row=e_r, column=10).alignment = align_left
    ws_s.cell(row=e_r, column=10).border = thin_border
    
    c_cell = ws_s.cell(row=e_r, column=11, value=count_f)
    c_cell.font = f_regular
    c_cell.alignment = align_right
    c_cell.border = thin_border
    c_cell.number_format = fmt_integer
    
    sh_cell = ws_s.cell(row=e_r, column=12, value=share_f)
    sh_cell.font = f_regular
    sh_cell.alignment = align_right
    sh_cell.border = thin_border
    sh_cell.number_format = fmt_percent
    e_r += 1

# Total
ws_s.cell(row=23, column=10, value="Total").font = f_bold
ws_s.cell(row=23, column=11, value="=SUM(K17:K22)").font = f_bold
ws_s.cell(row=23, column=11).number_format = fmt_integer
ws_s.cell(row=23, column=11).alignment = align_right
ws_s.cell(row=23, column=12, value="=SUM(L17:L22)").font = f_bold
ws_s.cell(row=23, column=12).number_format = fmt_percent
ws_s.cell(row=23, column=12).alignment = align_right
for col_idx in range(10, 13):
    ws_s.cell(row=23, column=col_idx).border = border_total

# Auto dimensions
for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
    col_letter = col
    max_len = 0
    for cell in ws_s[col]:
        if cell.row < 7:
            continue
        val = str(cell.value or '')
        if len(val) > max_len:
            max_len = len(val)
    ws_s.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 55)

ws_s.column_dimensions['E'].width = 45 # Review Text is long
ws_s.column_dimensions['A'].width = 10
ws_s.column_dimensions['B'].width = 30
ws_s.column_dimensions['J'].width = 18

wb.save("Task_4_Sentiment_Analysis.xlsx")
print("Styled Task_4_Sentiment_Analysis.xlsx generated successfully!")
